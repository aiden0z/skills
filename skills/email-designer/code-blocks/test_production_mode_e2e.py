"""End-to-end test for Production Mode: crystallize → produce."""

# Auto-install openpyxl if missing (required for these tests)
import subprocess
try:
    import openpyxl
except ImportError:
    print("openpyxl not found, installing...")
    subprocess.check_call(["pip", "install", "openpyxl", "-q"])
    import openpyxl

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))


def _import_excel_gen():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "excel_template_generator",
        os.path.join(os.path.dirname(__file__), "excel-template-generator.py"),
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def test_round_trip():
    """
    Simulate: design HTML → define sections → generate Excel → load Excel → fill template.
    Verify the filled content matches expected values.
    """
    mod = _import_excel_gen()

    sections = [
        {
            "name": "Header",
            "description": "邮件标题区",
            "columns": [
                {"name": "title", "description": "标题", "example": "第12期产品周报", "required": True},
                {"name": "date", "description": "日期", "example": "2026-03-30", "required": True},
            ],
            "single_row": True,
        },
        {
            "name": "Stats",
            "description": "核心指标",
            "columns": [
                {"name": "metric", "description": "指标名", "example": "新增用户", "required": True},
                {"name": "value", "description": "数值", "example": "1,234", "required": True},
                {"name": "trend", "description": "趋势", "example": "↑", "required": False,
                 "validation": ["↑", "↓", "—"]},
            ],
            "single_row": False,
        },
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, "template.xlsx")

        mod.generate_template(sections, xlsx_path)
        assert os.path.exists(xlsx_path)

        data = mod.load_excel_data(xlsx_path)

        assert "Header" in data
        assert len(data["Header"]) == 1
        assert data["Header"][0]["title"] == "第12期产品周报"
        assert data["Header"][0]["date"] == "2026-03-30"

        assert "Stats" in data
        assert len(data["Stats"]) == 3
        assert data["Stats"][0]["metric"] == "新增用户"
        assert data["Stats"][0]["value"] == "1,234"
        assert data["Stats"][0]["trend"] == "↑"

    print("PASS: test_round_trip")


def test_markup_in_excel():
    """Test that markup tags in Excel cells are processed correctly."""
    mod = _import_excel_gen()

    sections = [
        {
            "name": "Updates",
            "description": "更新内容",
            "columns": [
                {"name": "content", "description": "内容", "example": "<red>紧急</red>修复完成", "required": True},
            ],
            "single_row": False,
        },
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, "template.xlsx")
        mod.generate_template(sections, xlsx_path)
        data = mod.load_excel_data(xlsx_path)

        assert "Updates" in data
        content = data["Updates"][0]["content"]
        assert '<span style="color:#b91c1c">紧急</span>' in content
        assert "修复完成" in content

    print("PASS: test_markup_in_excel")


def test_empty_rows_skipped():
    """Test that empty rows in Excel are skipped during loading."""
    mod = _import_excel_gen()

    from openpyxl import load_workbook

    sections = [
        {
            "name": "Items",
            "description": "条目",
            "columns": [
                {"name": "name", "description": "名称", "example": "项目A", "required": True},
            ],
            "single_row": False,
        },
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, "template.xlsx")
        mod.generate_template(sections, xlsx_path)

        wb = load_workbook(xlsx_path)
        ws = wb["Items"]
        ws.cell(row=5, column=1, value="")
        ws.cell(row=6, column=1, value="有效数据")
        wb.save(xlsx_path)
        wb.close()

        data = mod.load_excel_data(xlsx_path)
        names = [row["name"] for row in data["Items"]]
        assert "" not in names
        assert "有效数据" in names

    print("PASS: test_empty_rows_skipped")


if __name__ == "__main__":
    test_round_trip()
    test_markup_in_excel()
    test_empty_rows_skipped()
    print("\nAll e2e tests passed!")
