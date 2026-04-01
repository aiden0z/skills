"""
Excel Template Generator — Generate & Load Excel Data Templates
================================================================
Two responsibilities:
1. Generate: Create an Excel template from section definitions (crystallization)
2. Load: Read data from a filled Excel template (production mode)

Layout: All email sections live in a single "邮件数据" sheet, laid out top-to-bottom
mirroring the email's visual flow. Section title rows act as anchors for parsing.
A separate "填写说明" sheet provides instructions and markup syntax reference.

This design follows the PoC quarterly report pattern (validated in production):
one data sheet with anchor-based section detection, resilient to row insertion/deletion.

Dependencies: openpyxl (auto-installed by deps-checker.py)
"""

import re
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel template generation. "
        "Install it with: pip install openpyxl"
    )


# ---------- Markup Processing ----------

MARKUP_COLORS = {
    "black": "#374151",
    "red": "#b91c1c",
    "green": "#059669",
    "orange": "#d97706",
    "blue": "#2563eb",
}


def process_markup(text: str) -> str:
    if not text or not isinstance(text, str):
        return ""
    html = text
    for color_name, color_hex in MARKUP_COLORS.items():
        pattern = rf"<({color_name})>(.+?)</\1>"
        html = re.sub(
            pattern,
            lambda m: f'<span style="color:{MARKUP_COLORS[m.group(1)]}">{m.group(2)}</span>',
            html, flags=re.DOTALL,
        )
    html = re.sub(
        r'<a\s+href=["\'](.+?)["\']\s*>(.+?)</a>',
        r'<a href="\1" style="color:#2563eb; text-decoration:underline;">\2</a>',
        html, flags=re.DOTALL,
    )
    html = html.replace("\r\n", "<br>").replace("\n", "<br>")
    return html


def process_cell_value(value) -> str:
    if value is None:
        return ""
    return process_markup(str(value).strip())


# ---------- Styles ----------

_FONT_FAMILY = "Microsoft YaHei"

SECTION_FONT = Font(name=_FONT_FAMILY, size=12, bold=True, color="FFFFFF")
SECTION_FILLS = [
    PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid"),
    PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid"),
    PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid"),
    PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid"),
    PatternFill(start_color="059669", end_color="059669", fill_type="solid"),
]
HEADER_FONT = Font(name=_FONT_FAMILY, size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
HINT_FONT = Font(name=_FONT_FAMILY, size=9, italic=True, color="6B7280")
DATA_FONT = Font(name=_FONT_FAMILY, size=10)
EXAMPLE_FONT = Font(name=_FONT_FAMILY, size=10, color="6B7280")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

# Section anchor marker format: [SectionName] in column A
_SECTION_MARKER_RE = re.compile(r'^\[(.+)\]$')


# ---------- Generate ----------

def generate_template(sections: List[Dict], output_path: str) -> str:
    """Generate an Excel template with all sections in a single data sheet.

    Args:
        sections: List of section definitions. Each section dict has:
            - name: str — Section name (English, matches SECTION comment in template.html)
            - description: str — Chinese description for display
            - columns: list of column defs (name, description, example, required, validation?)
            - single_row: bool — True for key-value sections, False for multi-row tables
        output_path: Where to save the .xlsx file.

    Returns:
        Absolute path to the generated file.

    Sheet layout (邮件数据):
        Each section is laid out vertically with:
        - Section title row: [Name] in col A, description in col B (styled, colored)
        - Column header row: field names starting from col B (dark header style)
        - Data rows: example values (gray italic), 1 row for single_row, 3 for multi-row
        - Empty separator row between sections
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 邮件数据 (all sections in one sheet)
    ws_data = wb.create_sheet("邮件数据")
    ws_data.sheet_properties.tabColor = "2563EB"
    _build_data_sheet(ws_data, sections)

    # Sheet 2: 填写说明
    ws_guide = wb.create_sheet("填写说明")
    _build_instructions_sheet(ws_guide, sections)

    wb.save(output_path)
    wb.close()
    return str(Path(output_path).resolve())


def _build_data_sheet(ws, sections: List[Dict]):
    """Build the single data sheet with all sections laid out vertically."""
    # Column A is narrow — holds section markers for parsing
    ws.column_dimensions["A"].width = 18

    row = 1

    for i, section in enumerate(sections):
        columns = section["columns"]
        fill = SECTION_FILLS[i % len(SECTION_FILLS)]
        single_row = section.get("single_row", False)
        hint = "单行数据" if single_row else "多行数据（可增减行）"
        num_cols = len(columns)

        # Determine how many spreadsheet columns this section spans (B onwards)
        last_col = 1 + num_cols  # col B = 2, so last = 1 + num_cols

        # ── Section title row ──
        # Column A: [SectionName] marker (for parser)
        marker_cell = ws.cell(row=row, column=1, value=f"[{section['name']}]")
        marker_cell.font = SECTION_FONT
        marker_cell.fill = fill
        marker_cell.alignment = CENTER_ALIGNMENT
        # Column B onwards: description + hint (merged)
        desc_text = f"{section['description']}（{hint}）"
        desc_cell = ws.cell(row=row, column=2, value=desc_text)
        desc_cell.font = SECTION_FONT
        desc_cell.fill = fill
        desc_cell.alignment = Alignment(horizontal="left", vertical="center")
        # Apply fill to all columns in this row
        for c in range(1, last_col + 1):
            ws.cell(row=row, column=c).fill = fill
        if last_col > 2:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
        ws.row_dimensions[row].height = 28
        row += 1

        # ── Column header row ──
        for col_idx, col_def in enumerate(columns):
            cell = ws.cell(row=row, column=2 + col_idx, value=col_def["name"])
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER
            # Add comment with description and example
            required_mark = "（必填）" if col_def.get("required") else "（可选）"
            cell.comment = Comment(
                f'{col_def["description"]}{required_mark}\n示例: {col_def.get("example", "")}',
                "Email Designer",
            )
        row += 1

        # ── Example data rows ──
        num_examples = 1 if single_row else 3
        for row_idx in range(num_examples):
            for col_idx, col_def in enumerate(columns):
                cell = ws.cell(row=row, column=2 + col_idx, value=col_def.get("example", ""))
                cell.font = EXAMPLE_FONT
                cell.alignment = WRAP_ALIGNMENT
                cell.border = THIN_BORDER
            row += 1

        # ── Add data validation (dropdowns) ──
        for col_idx, col_def in enumerate(columns):
            if "validation" in col_def and col_def["validation"]:
                dv = DataValidation(
                    type="list",
                    formula1='"' + ",".join(col_def["validation"]) + '"',
                    allow_blank=not col_def.get("required", False),
                )
                dv.error = f"请从下拉列表中选择{col_def['name']}的值"
                dv.errorTitle = "无效值"
                ws.add_data_validation(dv)
                col_letter = get_column_letter(2 + col_idx)
                # Apply from first data row to row+50 (enough room for user data)
                data_start = row - num_examples
                dv.add(f"{col_letter}{data_start}:{col_letter}{data_start + 50}")

        # ── Empty separator row ──
        row += 1

    # ── Set column widths based on content ──
    # Collect max widths across all sections
    col_widths = {}
    for section in sections:
        for col_idx, col_def in enumerate(section["columns"]):
            col_key = 2 + col_idx
            max_len = max(
                len(col_def["name"]),
                len(str(col_def.get("example", ""))),
                10,
            )
            width = min(max_len + 4, 50)
            col_widths[col_key] = max(col_widths.get(col_key, 0), width)

    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Freeze panes below the first section title
    ws.freeze_panes = "A2"


def _build_instructions_sheet(ws, sections: List[Dict]):
    """Build the instructions/guide sheet."""
    ws.sheet_properties.tabColor = "059669"

    ws["A1"] = "📖 邮件模板填写说明"
    ws["A1"].font = Font(name=_FONT_FAMILY, size=14, bold=True)
    ws.merge_cells("A1:C1")

    ws["A3"] = (
        "本 Excel 用于填写邮件内容数据。\"邮件数据\" Sheet 包含所有区域，"
        "自上而下的顺序对应邮件的视觉布局。"
    )
    ws["A3"].font = DATA_FONT
    ws.merge_cells("A3:C3")

    row = 5

    # Section overview table
    ws.cell(row=row, column=1, value="区域名称").font = Font(name=_FONT_FAMILY, size=10, bold=True)
    ws.cell(row=row, column=2, value="对应区域").font = Font(name=_FONT_FAMILY, size=10, bold=True)
    ws.cell(row=row, column=3, value="说明").font = Font(name=_FONT_FAMILY, size=10, bold=True)
    row += 1
    for section in sections:
        ws.cell(row=row, column=1, value=section["name"]).font = DATA_FONT
        ws.cell(row=row, column=2, value=section["description"]).font = DATA_FONT
        note = "单行数据" if section.get("single_row") else "多行数据（可增减行）"
        ws.cell(row=row, column=3, value=note).font = DATA_FONT
        row += 1
    row += 1

    # Field details per section
    for section in sections:
        ws.cell(row=row, column=1, value=f"📝 {section['name']} 字段说明").font = Font(
            name=_FONT_FAMILY, size=11, bold=True)
        row += 1
        for col_def in section["columns"]:
            required_mark = "（必填）" if col_def.get("required") else "（可选）"
            ws.cell(row=row, column=1, value=col_def["name"]).font = DATA_FONT
            ws.cell(row=row, column=2, value=f'{col_def["description"]}{required_mark}').font = DATA_FONT
            ws.cell(row=row, column=3, value=f'示例: {col_def.get("example", "")}').font = EXAMPLE_FONT
            row += 1
        row += 1

    # Markup syntax reference
    ws.cell(row=row, column=1, value="🎨 内容标记语法").font = Font(
        name=_FONT_FAMILY, size=11, bold=True)
    row += 1
    for label, example in [
        ("加粗", "<strong>文本</strong>"),
        ("红色文本", "<red>文本</red>（警告、紧急）"),
        ("绿色文本", "<green>文本</green>（成功、正向）"),
        ("蓝色文本", "<blue>文本</blue>（信息、链接）"),
        ("橙色文本", "<orange>文本</orange>（关注、待观察）"),
        ("链接", '<a href="https://example.com">点击查看</a>'),
    ]:
        ws.cell(row=row, column=1, value=label).font = DATA_FONT
        ws.cell(row=row, column=2, value=example).font = DATA_FONT
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 40


# ---------- Load ----------

def load_excel_data(excel_path: str) -> Dict[str, List[Dict]]:
    """Load data from a filled Excel template.

    Parses the "邮件数据" sheet using anchor-based section detection:
    1. Scans column A for [SectionName] markers
    2. For each section, reads column headers from the row after the marker
    3. Reads data rows until an empty row or the next section marker

    Returns:
        Dict mapping section names to lists of row dicts.
        Example: {"Header": [{"title": "周报", "date": "2026-03-30"}],
                  "Stats": [{"metric": "DAU", "value": "1234"}, ...]}

    Also supports legacy multi-sheet format for backward compatibility:
    if no "邮件数据" sheet exists, falls back to reading each sheet as a section.
    """
    wb = load_workbook(excel_path, data_only=True)

    # Prefer single-sheet format
    if "邮件数据" in wb.sheetnames:
        data = _load_single_sheet(wb["邮件数据"])
    else:
        # Legacy fallback: multi-sheet format
        data = _load_multi_sheet(wb)

    wb.close()
    return data


def _load_single_sheet(ws) -> Dict[str, List[Dict]]:
    """Parse the single data sheet using [SectionName] anchors in column A."""
    data = {}
    max_row = ws.max_row or 1

    # Step 1: Find all section anchor rows
    anchors = []  # [(row_number, section_name), ...]
    for row in range(1, max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            match = _SECTION_MARKER_RE.match(str(val).strip())
            if match:
                anchors.append((row, match.group(1)))

    # Step 2: Parse each section
    for idx, (anchor_row, section_name) in enumerate(anchors):
        # Boundary: next section anchor or end of sheet
        if idx + 1 < len(anchors):
            boundary = anchors[idx + 1][0]
        else:
            boundary = max_row + 1

        # Header row is immediately after the anchor
        header_row = anchor_row + 1
        if header_row >= boundary:
            continue

        # Read column headers from col B onwards
        headers = []
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val is None:
                break
            headers.append(str(val).strip())

        if not headers:
            continue

        # Read data rows
        section_data = []
        for row in range(header_row + 1, boundary):
            values = [ws.cell(row=row, column=2 + c).value for c in range(len(headers))]
            # Stop at empty row (all values None or blank)
            if all(v is None or str(v).strip() == "" for v in values):
                break
            row_dict = {}
            for header, value in zip(headers, values):
                row_dict[header] = process_cell_value(value)
            section_data.append(row_dict)

        data[section_name] = section_data

    return data


def _load_multi_sheet(wb) -> Dict[str, List[Dict]]:
    """Legacy loader: each sheet is a section (backward compatibility)."""
    data = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "填写说明":
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=False))
        if not rows:
            continue
        headers = [cell.value for cell in rows[0] if cell.value is not None]
        if not headers:
            continue
        sheet_data = []
        for row in rows[1:]:
            values = [cell.value for cell in row[:len(headers)]]
            if all(v is None or str(v).strip() == "" for v in values):
                continue
            row_dict = {}
            for header, value in zip(headers, values):
                row_dict[header] = process_cell_value(value)
            sheet_data.append(row_dict)
        data[sheet_name] = sheet_data
    return data
