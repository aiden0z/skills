"""Tests for Excel template generator."""

# Auto-install openpyxl if missing (required for these tests)
import subprocess
try:
    import openpyxl
except ImportError:
    print("openpyxl not found, installing...")
    subprocess.check_call(["pip", "install", "openpyxl", "-q"])
    import openpyxl

import os
import tempfile
import sys
sys.path.insert(0, os.path.dirname(__file__))


def _import():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "excel_template_generator",
        os.path.join(os.path.dirname(__file__), "excel-template-generator.py")
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def test_generate_template_creates_file():
    mod = _import()
    sections = [
        {
            "name": "Header",
            "description": "顶部标题栏",
            "columns": [
                {"name": "title", "description": "邮件标题", "example": "第12期周报", "required": True},
                {"name": "subtitle", "description": "副标题", "example": "2026年3月", "required": False},
                {"name": "date", "description": "日期", "example": "2026-03-30", "required": True},
            ],
            "single_row": True,
        },
        {
            "name": "Articles",
            "description": "文章列表区",
            "columns": [
                {"name": "title", "description": "文章标题", "example": "新功能发布", "required": True},
                {"name": "summary", "description": "摘要", "example": "本周发布了...", "required": True},
                {"name": "image_path", "description": "配图路径", "example": "banner.png", "required": False},
            ],
            "single_row": False,
        },
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "template.xlsx")
        mod.generate_template(sections, output_path)
        assert os.path.exists(output_path)
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        assert "填写说明" in wb.sheetnames
        assert "Header" in wb.sheetnames
        assert "Articles" in wb.sheetnames
        wb.close()


def test_load_excel_data():
    mod = _import()
    sections = [
        {
            "name": "Stats",
            "description": "数据指标",
            "columns": [
                {"name": "metric", "description": "指标名", "example": "新增用户", "required": True},
                {"name": "value", "description": "值", "example": "1,234", "required": True},
            ],
            "single_row": False,
        },
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "template.xlsx")
        mod.generate_template(sections, output_path)
        data = mod.load_excel_data(output_path)
        assert "Stats" in data
        assert len(data["Stats"]) >= 1
        assert "metric" in data["Stats"][0]
        assert "value" in data["Stats"][0]


def test_markup_processing():
    mod = _import()
    assert mod.process_markup("<red>紧急</red>") == '<span style="color:#b91c1c">紧急</span>'
    assert mod.process_markup("<strong>重要</strong>") == "<strong>重要</strong>"
    assert mod.process_markup("<green>完成</green>") == '<span style="color:#059669">完成</span>'
    assert mod.process_markup("普通文本") == "普通文本"
    assert mod.process_markup("<blue>信息</blue>") == '<span style="color:#2563eb">信息</span>'
    assert mod.process_markup("<orange>警告</orange>") == '<span style="color:#d97706">警告</span>'


if __name__ == "__main__":
    test_generate_template_creates_file()
    print("PASS: test_generate_template_creates_file")
    test_load_excel_data()
    print("PASS: test_load_excel_data")
    test_markup_processing()
    print("PASS: test_markup_processing")
    print("\nAll tests passed!")
